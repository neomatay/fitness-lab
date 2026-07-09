#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从「好人松松 健身Excel套表」提取结构化数据 → JSON
产物: src/data/foods.json, exercises.json, qa.json
"""
import json
import os
import re
import openpyxl

# Excel 文件路径（与脚本同级的上级目录）
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
EXCEL = os.path.join(PROJECT_DIR, "..", "【可任意分享】健身Excel超级套表（作者：B站好人松松）26年4月最新版.xlsx")
EXCEL = os.path.abspath(EXCEL)
OUT_DIR = os.path.join(PROJECT_DIR, "src", "data")
os.makedirs(OUT_DIR, exist_ok=True)

wb = openpyxl.load_workbook(EXCEL, read_only=False, data_only=True)


def cell(ws, coord):
    v = ws[coord].value
    if v is None:
        return None
    return str(v).strip()


# ============================================================
# 1. 食物营养库 (表19)
# ============================================================
def extract_foods():
    ws = wb["19日常食物营养率"]
    foods = []
    section = None        # 碳水 / 蛋白质 / 脂肪
    category = None       # 大类（米类主食等）

    # 表头行: R2=B2碳水, R93=B93蛋白质, R117=B117脂肪
    # 列: B=大类 C=食物 D=碳水率/蛋白质率 E=GI/定位 F=讲解
    for r in range(1, ws.max_row + 1):
        b = cell(ws, f"B{r}")
        c = cell(ws, f"C{r}")
        d = ws[f"D{r}"].value
        e = cell(ws, f"E{r}")
        f = cell(ws, f"F{r}")

        if b == "碳水":
            section = "碳水"
            continue
        if b == "蛋白质":
            section = "蛋白质"
            continue
        if b == "脂肪":
            section = "脂肪"
            continue

        # 脂肪部分整体是文字说明，不存为食物条目
        if section == "脂肪":
            continue

        # 大类行：B列是大类名（米类主食/麦类主食等），表头"大类"跳过
        # 注意：大类行 C 列通常也有食物名，所以判断大类后不能 continue，要继续处理该行食物
        if b and b != "大类" and section in ("碳水", "蛋白质"):
            # 判断是否是大类名：不含数字、不是表头、且 C 列该食物属于新大类
            # 大类名特征：纯中文，无括号数字，且与"食物"表头不同
            is_category = (b != "食物" and not re.search(r"\d", b)
                           and len(b) <= 6 and "率" not in b)
            if is_category:
                category = b
                # 不 continue，继续处理同行的 C 列食物

        # 食物条目行：C列有食物名
        if c and c != "食物":
            item = {
                "category": category or "",
                "name": c,
                "note": f or "",
            }
            # 率值处理（可能是小数，也可能是 "6g/个" 这种文本）
            if d is not None and d != "":
                ds = str(d).strip()
                # 尝试解析为数字
                try:
                    rate = float(ds)
                    if section == "碳水":
                        item["carbRate"] = rate
                    else:
                        item["proteinRate"] = rate
                except ValueError:
                    # 非数字（如 "6g/个" "10g/250ml盒"），原样存
                    if section == "碳水":
                        item["carbRateText"] = ds
                    else:
                        item["proteinRateText"] = ds
            if section == "碳水":
                # GI 在 E 列
                if e:
                    item["gi"] = e
            else:
                # 蛋白质的 E 列是"定位"（主要来源/早饭来源等）
                if e:
                    item["role"] = e
            foods.append(item)

    return foods


# ============================================================
# 2. 训练动作库 (表20-23)
# ============================================================
def extract_exercises():
    sheets = {
        "三分化": "20训练计划-健身房三分化",
        "四分化肩单练": "21训练计划-健身房四分化计划（肩单练版）",
        "四分化臂单练": "22训练计划-健身房四分化计划（手臂单练版）",
        "居家": "23训练计划-居家健身",
    }
    exercises = []

    # 判断是否为说明性长文本（非部位名）：含句号/换行/说明词/超长
    def is_description(text):
        if not text:
            return False
        if len(text) > 12:
            return True
        if any(kw in text for kw in ["是三个", "建议", "可以", "按以下", "循环", "男第", "女第"]):
            return True
        return False

    # 清洗部位名：去掉换行后的说明（如"腹\n男性要练 女性偶尔练下" → "腹"）
    def clean_part(text):
        if not text:
            return text
        return text.split("\n")[0].strip()

    # 清洗 day 名：去掉括注（如"（请看下表左侧红字选动作）"）
    def clean_day(text):
        if not text:
            return text
        # 去掉中文括号及内容
        text = re.sub(r"（[^（）]*）", "", text)
        text = re.sub(r"\([^()]*\)", "", text)
        return text.strip()

    for plan, sheet_name in sheets.items():
        ws = wb[sheet_name]
        day = None
        part = None
        group = None

        for r in range(1, ws.max_row + 1):
            b = cell(ws, f"B{r}")
            c = cell(ws, f"C{r}")
            d = cell(ws, f"D{r}")
            e = cell(ws, f"E{r}")
            f = cell(ws, f"F{r}")

            # Day 标记 (Day1：背+肩后束+肱二头 等)
            if b and re.match(r"^Day\s*\d", b):
                day = clean_day(b)
                part = None
                group = None
                continue

            # 部位行：B列有值（如"背"、"胸"），且C列是组数说明
            if b and b not in ("部位", "知识准备", "训练频率", "训练组数",
                               "组间休息", "配重选择", "是否力竭", "女性注意"):
                bclean = b.strip()
                # 部位行通常 B=部位名, C="选1-2个动作 总共6-8组" 这类
                # 有C列组数的，清洗换行说明后作为部位（如"腹\n男性要练" → "腹"）
                if c and ("组" in c or "选" in c):
                    part = clean_part(bclean)
                    group = c
                    continue
                # 无C列组数：跳过说明性长文本（不是部位名）
                if is_description(bclean):
                    continue
                # 仅部位名无组数（少见）
                if not c and part != bclean:
                    part = clean_part(bclean)
                    continue

            # 动作行：D列有动作名
            if d and d not in ("动作",):
                # part 已在源头清洗；若仍是说明性文本则置空避免污染
                final_part = part if (part and not is_description(part)) else ""
                ex = {
                    "plan": plan,
                    "day": day or "",
                    "part": final_part,
                    "group": group or "",
                    "name": d,
                }
                if e and e not in ("肩关节", "膝关节"):
                    ex["shoulder"] = e
                if f and f not in ("肘关节", "髋关节"):
                    ex["elbow"] = f
                exercises.append(ex)

    return exercises


# ============================================================
# 3. 问答库 (表17-18)
# ============================================================
def extract_qa():
    qa = []

    for qtype, sheet_name in [("减脂", "17减脂-问答汇总"), ("增肌", "18增肌-问答汇总")]:
        ws = wb[sheet_name]
        rows = []
        for r in range(1, ws.max_row + 1):
            b = cell(ws, f"B{r}")
            if b:
                rows.append(b)

        # 解析：目录是"1.xxx?"样式，正文是问题+答案交替的大段文本
        # 策略：找形如 "数字.问题？" 的行作为问题，下一个非空行作为答案（可能跨多行合并）
        i = 0
        # 先跳过标题区，找到第一个问题
        questions = []
        current_q = None
        current_a_lines = []
        collecting = False

        for line in rows:
            # 匹配问题标题 " 1.执行了2周但体重不掉该怎么调整？ "
            m = re.match(r"^\s*(\d+)\.\s*(.+[？?])\s*$", line)
            # 目录行（无答案紧随）形如 "1.执行了..." 在目录区
            # 正文行也形如 " 1.执行了2周...？ "

            if m and not collecting:
                # 可能是目录项，先记录，等遇到正文重复时开始收集
                current_q = m.group(2).strip()
                current_a_lines = []
                collecting = True
                continue

            if collecting:
                # 如果遇到下一个问题，保存上一个
                m2 = re.match(r"^\s*(\d+)\.\s*(.+[？?])\s*$", line)
                if m2:
                    # 保存前一个
                    if current_q and current_a_lines:
                        qa.append({
                            "type": qtype,
                            "q": current_q,
                            "a": "\n".join(current_a_lines).strip()
                        })
                    current_q = m2.group(2).strip()
                    current_a_lines = []
                else:
                    current_a_lines.append(line)

        # 最后一个
        if current_q and current_a_lines:
            qa.append({
                "type": qtype,
                "q": current_q,
                "a": "\n".join(current_a_lines).strip()
            })

    # 去重：同一问题可能因目录+正文出现两次，保留答案更长的
    seen = {}
    for item in qa:
        key = (item["type"], item["q"])
        if key not in seen or len(item["a"]) > len(seen[key]["a"]):
            seen[key] = item
    return list(seen.values())


# ============================================================
# 主流程
# ============================================================
if __name__ == "__main__":
    print("提取食物库...")
    foods = extract_foods()
    print(f"  → {len(foods)} 条食物")

    print("提取训练动作...")
    exercises = extract_exercises()
    print(f"  → {len(exercises)} 个动作")

    print("提取问答...")
    qa = extract_qa()
    print(f"  → {len(qa)} 条问答")

    with open(os.path.join(OUT_DIR, "foods.json"), "w", encoding="utf-8") as fp:
        json.dump(foods, fp, ensure_ascii=False, indent=2)
    with open(os.path.join(OUT_DIR, "exercises.json"), "w", encoding="utf-8") as fp:
        json.dump(exercises, fp, ensure_ascii=False, indent=2)
    with open(os.path.join(OUT_DIR, "qa.json"), "w", encoding="utf-8") as fp:
        json.dump(qa, fp, ensure_ascii=False, indent=2)

    print(f"\n完成！数据已写入 {OUT_DIR}")
    print(f"  foods.json: {len(foods)} 条")
    print(f"  exercises.json: {len(exercises)} 条")
    print(f"  qa.json: {len(qa)} 条")

    wb.close()
